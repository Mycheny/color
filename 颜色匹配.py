# -*- coding: utf-8 -*- 
# @Time 2020/3/5 23:45
# @Author wcy

from gevent import monkey

monkey.patch_all()
import gevent
from haishoku.haishoku import Haishoku
import math
from colorsys import rgb_to_hsv
import os
from collections import OrderedDict
import pandas as pd
import time

from openpyxl import Workbook
from openpyxl.styles import PatternFill, fills

colors = dict((
    ((255, 182, 193), "浅粉色"),
    ((255, 192, 203), "粉红"),
    ((220, 20, 60), "猩红"),
    ((255, 240, 245), "脸红的淡紫色"),
    ((219, 112, 147), "苍白的紫罗兰红色"),
    ((255, 105, 180), "热情的粉红"),
    ((255, 20, 147), "深粉色"),
    ((199, 21, 133), "适中的紫罗兰红色"),
    ((218, 112, 214), "兰花的紫色"),
    ((216, 191, 216), "蓟"),
    ((221, 160, 221), "李子"),
    ((238, 130, 238), "紫罗兰"),
    ((255, 0, 255), "洋红"),
    ((139, 0, 139), "深洋红色"),
    ((128, 0, 128), "紫色"),
    ((186, 85, 211), "适中的兰花紫"),
    ((148, 0, 211), "深紫罗兰色"),
    ((153, 50, 204), "深兰花紫"),
    ((75, 0, 130), "靛青"),
    ((138, 43, 226), "深紫罗兰的蓝色"),
    ((147, 112, 219), "适中的紫色"),
    ((123, 104, 238), "适中的板岩暗蓝灰色"),
    ((106, 90, 205), "板岩暗蓝灰色"),
    ((72, 61, 139), "深岩暗蓝灰色"),
    ((230, 230, 250), "薰衣草花的淡紫色"),
    ((248, 248, 255), "幽灵的白色"),
    ((0, 0, 255), "纯蓝"),
    ((0, 0, 205), "适中的蓝色"),
    ((25, 25, 112), "午夜的蓝色"),
    ((0, 0, 139), "深蓝色"),
    ((0, 0, 128), "海军蓝"),
    ((65, 105, 225), "宝蓝"),
    ((100, 149, 237), "矢车菊的蓝色"),
    ((176, 196, 222), "淡钢蓝"),
    ((119, 136, 153), "浅石板灰"),
    ((112, 128, 144), "石板灰"),
    ((30, 144, 255), "道奇蓝"),
    ((240, 248, 255), "爱丽丝蓝"),
    ((70, 130, 180), "钢蓝"),
    ((135, 206, 250), "淡蓝色"),
    ((135, 206, 235), "天蓝色"),
    ((0, 191, 255), "深天蓝"),
    ((173, 216, 230), "淡蓝"),
    ((176, 224, 230), "火药蓝"),
    ((95, 158, 160), "军校蓝"),
    ((240, 255, 255), "蔚蓝色"),
    ((225, 255, 255), "淡青色"),
    ((175, 238, 238), "苍白的绿宝石"),
    ((0, 255, 255), "水绿色"),
    ((0, 206, 209), "深绿宝石"),
    ((47, 79, 79), "深石板灰"),
    ((0, 139, 139), "深青色"),
    ((0, 128, 128), "水鸭色"),
    ((72, 209, 204), "适中的绿宝石"),
    ((32, 178, 170), "浅海洋绿"),
    ((64, 224, 208), "绿宝石"),
    ((127, 255, 170), "绿玉"),
    ((0, 250, 154), "适中的碧绿色"),
    ((245, 255, 250), "适中的春天的绿色"),
    ((0, 255, 127), "薄荷奶油"),
    ((60, 179, 113), "春天的绿色"),
    ((46, 139, 87), "海洋绿"),
    ((240, 255, 0), "蜂蜜"),
    ((144, 238, 144), "淡绿色"),
    ((152, 251, 152), "苍白的绿色"),
    ((143, 188, 143), "深海洋绿"),
    ((50, 205, 50), "酸橙绿"),
    ((0, 255, 0), "酸橙色"),
    ((34, 139, 34), "森林绿"),
    ((0, 128, 0), "纯绿"),
    ((0, 100, 0), "深绿色"),
    ((127, 255, 0), "查特酒绿"),
    ((124, 252, 0), "草坪绿"),
    ((173, 255, 47), "绿黄色"),
    ((85, 107, 47), "橄榄土褐色"),
    ((107, 142, 35), "米色（浅褐色）"),
    ((250, 250, 210), "浅秋麒麟黄"),
    ((255, 255, 240), "象牙色"),
    ((255, 255, 224), "浅黄色"),
    ((255, 255, 0), "纯黄"),
    ((128, 128, 0), "橄榄"),
    ((189, 183, 107), "深卡其布"),
    ((255, 250, 205), "柠檬薄纱"),
    ((238, 232, 170), "灰秋麒麟"),
    ((240, 230, 140), "卡其布"),
    ((255, 215, 0), "金"),
    ((255, 248, 220), "玉米色"),
    ((218, 165, 32), "秋麒麟"),
    ((255, 250, 240), "花的白色"),
    ((253, 245, 230), "老饰带"),
    ((245, 222, 179), "小麦色"),
    ((255, 228, 181), "鹿皮鞋"),
    ((255, 165, 0), "橙色"),
    ((255, 239, 213), "番木瓜"),
    ((255, 235, 205), "漂白的杏仁"),
    ((255, 222, 173), "Navajo白"),
    ((250, 235, 215), "古代的白色"),
    ((210, 180, 140), "晒黑"),
    ((222, 184, 135), "结实的树"),
    ((255, 228, 196), "（浓汤）乳脂，番茄等"),
    ((255, 140, 0), "深橙色"),
    ((250, 240, 230), "亚麻布"),
    ((205, 133, 63), "秘鲁"),
    ((255, 218, 185), "桃色"),
    ((244, 164, 96), "沙棕色"),
    ((210, 105, 30), "巧克力"),
    ((139, 69, 19), "马鞍棕色"),
    ((255, 245, 238), "海贝壳"),
    ((160, 82, 45), "黄土赭色"),
    ((255, 160, 122), "浅鲜肉（鲑鱼）色"),
    ((255, 127, 80), "珊瑚"),
    ((255, 69, 0), "橙红色"),
    ((233, 150, 122), "深鲜肉（鲑鱼）色"),
    ((255, 99, 71), "番茄"),
    ((255, 228, 225), "薄雾玫瑰"),
    ((250, 128, 114), "鲜肉（鲑鱼）色"),
    ((255, 250, 250), "雪"),
    ((240, 128, 128), "淡珊瑚色"),
    ((188, 143, 143), "玫瑰棕色"),
    ((205, 92, 92), "印度红"),
    ((255, 0, 0), "纯红"),
    ((165, 42, 42), "棕色"),
    ((178, 34, 34), "耐火砖"),
    ((139, 0, 0), "深红色"),
    ((128, 0, 0), "栗色"),
    ((255, 255, 255), "纯白"),
    ((245, 245, 245), "白烟"),
    ((220, 220, 220), "Gainsboro"),
    ((211, 211, 211), "浅灰色"),
    ((192, 192, 192), "银白色"),
    ((169, 169, 169), "深灰色"),
    ((128, 128, 128), "灰色"),
    ((105, 105, 105), "暗淡的灰色"),
    ((0, 0, 0), "纯黑")
))

print()

def to_hsv(color):
    """ converts color tuples to floats and then to hsv """
    return rgb_to_hsv(*[x / 255.0 for x in color])  # rgb_to_hsv wants floats!


def color_dist(c1, c2):
    """ returns the squared euklidian distance between two color vectors in hsv space """
    return sum((a - b) ** 2 for a, b in zip(to_hsv(c1), to_hsv(c2)))


def min_color_diff(color_to_match, colors):
    """ returns the `(distance, color_name)` with the minimal distance to `colors`"""
    return min(  # overal best is the best match to any color:
        (color_dist(color_to_match, test), colors[test])  # (distance to `test` color, color name)
        for test in colors)


def rgb2hex(rgbcolor):
    r, g, b = rgbcolor
    color = "#"
    color += str(hex(r)).replace('x', '0')[-2:]
    color += str(hex(g)).replace('x', '0')[-2:]
    color += str(hex(b)).replace('x', '0')[-2:]
    return color


# 新建一个新的工作表（未保存）。
wb = Workbook()
# 保存文件，若加载路径与保存的路径一致将会被覆盖
my_sheet = wb.worksheets[0]
row_title = ["图片路径名称","主要图片颜色","主要图片颜色hex","主要图片颜色占比","次要图片颜色","次要图片颜色hex","次要图片颜色占比","次次要图片颜色","次次要图片颜色hex","次次要图片颜色占比"]
my_sheet.append(row_title)


def get_colorname(index,path):
    single_data = OrderedDict()
    try:
        haishoku = Haishoku.loadHaishoku(path)
        # Haishoku.showDominant(path)
        # single_data['图片路径名称'] = path
        palette = haishoku.palette
        main_color = palette[0][1]
        main_color_pct = palette[0][0]
        mian_colorname = min_color_diff(main_color, colors)[1]
        tmp_main_color = rgb2hex(main_color)[1:]
        # single_data['主要图片颜色'] = mian_colorname
        # single_data['主要图片颜色rgb'] = rgb2hex(main_color)
        # single_data['主要图片颜色占比'] = main_color_pct
        second_color = palette[1][1]
        second_color_pct = palette[1][0]
        second_colorname = min_color_diff(second_color, colors)[1]
        tmp_second_color = rgb2hex(second_color)[1:]
        # single_data['次要图片颜色'] = second_colorname
        # single_data['次要图片颜色rgb'] = rgb2hex(second_color)
        # single_data['次要图片颜色占比'] = second_color_pct
        thred_color = palette[2][1]
        thred_color_pct = palette[2][0]
        thred_colorname = min_color_diff(thred_color, colors)[1]
        tmp_thred_color = rgb2hex(thred_color)[1:]
        # single_data['次次要图片颜色'] = thred_colorname
        # single_data['次次要图片颜色rgb'] = rgb2hex(thred_color)
        # single_data['次次要图片颜色占比'] = thred_color_pct
        row_line = [path,mian_colorname,"#"+tmp_main_color,main_color_pct,second_colorname,"#"+tmp_second_color,second_color_pct,thred_colorname,"#"+tmp_thred_color,thred_color_pct]
        my_sheet.append(row_line)
        my_sheet["B"+str(2+int(index))].fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=rgb2hex(list(colors.keys())[list(colors.values()).index(mian_colorname)])[1:], bgColor=tmp_main_color)
        my_sheet["C"+str(2+int(index))].fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=tmp_main_color, bgColor=tmp_main_color)
        my_sheet["E"+str(2+int(index))].fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=rgb2hex(list(colors.keys())[list(colors.values()).index(second_colorname)])[1:], bgColor=tmp_main_color)
        my_sheet["F"+str(2+int(index))].fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=tmp_second_color, bgColor=tmp_second_color)
        my_sheet["H"+str(2+int(index))].fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=rgb2hex(list(colors.keys())[list(colors.values()).index(thred_colorname)])[1:], bgColor=tmp_main_color)
        my_sheet["I"+str(2+int(index))].fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=tmp_thred_color, bgColor=tmp_thred_color)
        print(index+1)
        # print(path,'主要颜色是：'+mian_colorname,'主要颜色占比：'+str(main_color_pct*100)+'%',' 次要颜色是：'+second_colorname,'次要颜色占比：'+str(second_color_pct*100)+'%',' 次次要颜色是：'+thred_colorname,'次次要颜色占比：'+str(thred_color_pct*100)+'%')
    except Exception as e:
        print('错误-->', path, e)

# root = "E:\\PycharmProjects\\服饰颜色识别\\image"
root = "C:\\Users\\Xiaoi\\Desktop\\没有标注的颜色4500张图片\\没有标注的颜色4500张图片"
pic_names = os.listdir(root)[:500]
# all_abs_picname = (os.path.join('./picture', name) for name in pic_names if name != '.DS_Store' and len(name)>4)
all_abs_picname = (os.path.join(root, name) for name in pic_names)

# greenlets = [gevent.spawn(get_colorname, index, path) for index, path in enumerate(all_abs_picname)]
# gevent.joinall(greenlets)

a = [get_colorname(index, path) for index, path in enumerate(all_abs_picname)]


if not os.path.exists('图片颜色导出'):
    os.mkdir(os.path.join(os.getcwd(), '图片颜色导出'))


wb.save(r'图片颜色导出' + '/' + time.strftime("%Y%m%d%H%M") + '.xlsx')
print('done')